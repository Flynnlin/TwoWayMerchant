import markdown
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import load_workbook
from merchant.forms import PlatformCategoryForm, ProductSourceForm, XhsOrderForm, XhsEditOrderForm, WikiModelForm
from merchant.models import PlatformCategory, ProductSource, XhsOrder, Wiki
from merchant.utils import uuidStr
from merchant.utils.pagenation import Pagenation


def platform_list_view(request):
    platforms= PlatformCategory.objects.all()
    return render(request, 'platform/platform_list.html', {"platforms":platforms})

def platform_add_view(request):
    if request.method == 'POST':
        form = PlatformCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/platform/")
        else:
            return render(request, 'platform/platform_add.html', {"form": form})
    form = PlatformCategoryForm()
    return render(request, 'platform/platform_add.html', {"form":form})

def platform_edit_view(request, pk):
    platform = PlatformCategory.objects.get(id=pk)
    if request.method == 'POST':
        form = PlatformCategoryForm(request.POST, instance=platform)
        if form.is_valid():
            form.save()
            return redirect("/platform/")
        else:
            return render(request, 'platform/platform_edit.html', {"form": form})

    form = PlatformCategoryForm(instance=platform)
    return render(request, 'platform/platform_edit.html', {"form":form})

def platform_delete_view(request, pk):
    delete = PlatformCategory.objects.get(id=pk).delete()
    return redirect("/platform/")


def merchandise_list_view(request):
    q_dict = {}
    q_uuid = request.GET.get('uuid')
    q_title = request.GET.get('title')
    if q_uuid :
        q_dict["uuid"] = q_uuid
    if q_title:
        q_dict["title__contains"] = q_title
    merchandises = ProductSource.objects.filter(**q_dict).order_by('-id')
    page_obj = Pagenation(merchandises, request)
    page_queryset = page_obj.page_queryset
    page_string = page_obj.html()
    return render(request, 'merchandise/merchandise_list.html', {"form":page_queryset,'page_string':page_string})
def merchandise_add_view(request):
    if request.method == 'POST':
        form = ProductSourceForm(request.POST)
        if form.is_valid():
            form.instance.uuid = uuidStr.generate_order_number()
            form.save()
            Wiki.objects.create(product_id=form.instance.id,title=form.cleaned_data['title'])
            return redirect("/merchandise/")
        else:
            return render(request, 'merchandise/merchandise_add.html',{"form":form})
    form = ProductSourceForm()
    return render(request, 'merchandise/merchandise_add.html', {"form":form})
def merchandise_edit_view(request, pk):
    merchandise=ProductSource.objects.get(id=pk)
    if request.method == 'POST':
        form = ProductSourceForm(request.POST, instance=merchandise)
        if form.is_valid():
            form.save()
            return redirect("/merchandise/")
        else:
            return render(request, 'merchandise/merchandise_edit.html', {"form": form})
    form = ProductSourceForm(instance=merchandise)
    return render(request, 'merchandise/merchandise_edit.html', {"form":form})
def merchandise_delete_view(request, pk):
    delete=ProductSource.objects.get(id=pk).delete()
    return redirect("/merchandise/")


def merchandise_wiki_view(request, pk):
    wiki= Wiki.objects.get(product_id=pk)
    if wiki.content:
        extensions = [
            'extra',
            'toc',
            'tables',
            'codehilite',
            'fenced_code',
            'footnotes',
            'admonition',
            'attr_list',
            'def_list',
            'abbr',
            'md_in_html',
            'meta',
            'nl2br',
            'sane_lists',
            'smarty',
            'wikilinks',
        ]
        wiki.content=markdown.markdown(wiki.content,extensions=extensions)
        print(wiki.content)
        wiki.content = wiki.content.replace('<img', '<img class="markdown-img"')
    return render(request, 'merchandise/wiki_show.html',{'wiki':wiki})

def merchandise_wiki_edit_view(request,pk):
    wiki = Wiki.objects.get(id=pk)
    if request.method == 'POST':
        form = WikiModelForm(request.POST, instance=wiki)
        if form.is_valid():
            form.save()
            return redirect(reverse("merchandise_wiki_view",args=[wiki.product.id]))
        else:
            return render(request, 'merchandise/wiki_edit.html', {"form":form})

    form = WikiModelForm(instance=wiki)
    return render(request, 'merchandise/wiki_edit.html', {"form":form})




def xhs_order_list_view(request):
    q_dict = {}
    q_uuid = request.GET.get('uuid')
    q_orderId = request.GET.get('orderId')
    q_date = request.GET.get('date')
    if q_uuid:
        q_dict["uuid"] = q_uuid
    if q_orderId:
        q_dict["orderId"] = q_orderId
    if q_date:
        q_dict["create_datetime__contains"] = q_date

    order=XhsOrder.objects.filter(**q_dict).order_by('create_datetime')
    page_obj = Pagenation(order, request)
    page_queryset = page_obj.page_queryset
    page_string = page_obj.html()
    return render(request, 'order/order_list.html', {"form":page_queryset,'page_string':page_string})

def xhs_order_detail_view(request, pk):
    order=XhsOrder.objects.get(id=pk)
    product = ProductSource.objects.get(uuid=order.product_source)
    return render(request, 'order/order_detail.html', {"order":order,"product":product})

def xhs_order_create_view(request):
    if request.method == 'POST':
        form = XhsOrderForm(request.POST)
        if form.is_valid():
            product= ProductSource.objects.get(uuid=form.cleaned_data['product_source'])
            #计算收益
            inaccount=product.output_price-product.input_price# 单个
            inaccount=inaccount*form.cleaned_data['amount'] #
            inaccount=inaccount-product.postage # 减去邮费
            print(inaccount)
            form.instance.platform=product.platform
            form.instance.inaccount=inaccount
            form.save()
            return redirect("/order/")
        else:
            return render(request, 'order/order_add.html', {"form": form})

    form = XhsOrderForm()
    return render(request, 'order/order_add.html',{"form":form})
def xhs_order_update_view(request, pk):
    order = XhsOrder.objects.get(id=pk)
    if request.method == 'POST':
        form = XhsEditOrderForm(data=request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect("/order/")
        else:
            return render(request, 'order/order_edit.html', {"form": form})

    form = XhsEditOrderForm(instance=order)
    return render(request, 'order/order_edit.html',{"form":form})
def xhs_order_delete_view(request, pk):
    delete = XhsOrder.objects.get(id=pk).delete()
    return redirect("/order/")



def merchandise_mutiladd(request):
    if request.method == 'POST':
        # 获取文件对象
        file_obj = request.FILES.get("file")
        Fname = file_obj.name
        # 打开读取Excel
        wb = load_workbook(file_obj)
        sheet = wb.active  # 获取当前活动的工作表
        # 循环每一行数据，跳过第一行（标题行）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 根据Excel列顺序和内容，逐行读取Excel数据，并根据数据模型创建商品对象
            platform_name, url, title, input_price, output_price, min_amount, postage= row
            # 检查平台是否存在
            platform_obj, created = PlatformCategory.objects.get_or_create(name=platform_name)
            # 如果需要根据平台名进行检查
            if not created:
                # 平台已存在，可以执行相应操作，例如创建商品对象
                ProductSource.objects.create(
                    platform=platform_obj,
                    url=url,
                    title=title,
                    input_price=input_price,
                    output_price=output_price,
                    min_amount=min_amount,
                    postage=postage,
                    uuid=uuidStr.generate_order_number()
                )
            else:
                # 平台不存在，可以进行相应处理，例如记录日志或返回错误信息
                pass
        return redirect("/merchandise/list")
    return render(request, 'merchandise/mutiladd.html')


# def xhs_order_mutiladd(request):
#     if request.method == 'POST':
#         # 获取文件对象
#         file_obj = request.FILES.get("file")
#         Fname = file_obj.name
#         # 打开读取Excel
#         wb = load_workbook(file_obj)
#         sheet = wb.active  # 获取当前活动的工作表
#         # 循环每一行数据，跳过第一行（标题行）
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             # 根据Excel列顺序和内容，逐行读取Excel数据，并根据数据模型创建订单对象
#             orderId, platform_name, platform_orderId, inaccount, product_source_uuid, amount, _ = row
#             # 检查平台是否存在
#             try:
#                 platform = PlatformCategory.objects.get(name=platform_name)
#             except PlatformCategory.DoesNotExist:
#                 return render(request, 'error.html', {'error_message': f'平台"{platform_name}"不存在'})
#             # 创建订单对象并保存到数据库中
#             XhsOrder.objects.create(
#                 orderId=orderId,
#                 platform=platform,
#                 platform_orderId=platform_orderId,
#                 inaccount=inaccount,
#                 product_source=product_source_uuid,
#                 amount=amount
#             )
#         return redirect("/xhs_order/list")
#     return render(request, 'xhs_order/xhs_order_mutiladd.html')
